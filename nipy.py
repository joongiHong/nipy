'''

[NiPy]

대부분의 코드에는 주석이 포함되어 있습니다.
Github 문서와 주석을 참고하여 필요한 코드만 살려
경량화 하시기 바랍니다.

'''

# ~~ 0. 임포트 부분 ~~
import requests  # (필수) 나이스 서버와 통신용
import json  # (필수) api 사용시 파싱용
from bs4 import BeautifulSoup  # (필수) 나이스 페이지 파싱용

import openpyxl  # (선택) 엑셀 저장용
import csv  # (선택) csv 저장용


# ~~ 1. 학교 코드를 불러오는 api ~~

class Scode:  # Scode 클래스 생성
    def __init__(self, name, city):  # 학교 이름과 위치를 받아옴
        city_dict = {"서울": "1100000000", "부산": "2600000000", "대구": "2700000000",
                     "인천": "2800000000", "광주": "2900000000", "대전": "3000000000",
                     "울산": "3100000000", "세종": "3600000000", "경기": "4100000000",
                     "강원": "4200000000", "충북": "4300000000", "충남": "4400000000",
                     "전북": "4500000000", "전남": "4600000000", "경북": "4700000000",
                     "경남": "4800000000", "제주": "5000000000"}  # 지역 목록

        self.city = city_dict.get(city, "")  # 지역 코드로 변환
        self.name = name  # 학교 이름 저장

    def codefind(self, kind):  # 실질적으로 코드를 반환하는 부분
        url = "https://www.schoolinfo.go.kr/ei/ss/Pneiss_a01_l0.do"  # 학교알리미 코드 불러오는 주소
        para = {
            "HG_CO": "",
            "SEARCH_KIND": "",
            "HG_JONGRYU_GB": "",
            "GS_HANGMOK_CD": "",
            "GU_GUN_CODE": "",
            "GUGUN_CODE": "",
            "SIDO_CODE": self.city,
            "SRC_HG_NM": self.name
        }  # post 파라미터
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8"
        }  # 헤더 데이터

        re = requests.post(url, data=para, headers=headers)  # 서버 접속
        rejs = json.loads(re.text)  # json 데이터 불러오기
        reco = re.status_code  # 응답코드 저장

        if int(reco) != 200:
            return("SERVER ERROR")  # 접속 여부 확인

        elementary = rejs['schoolList02']  # 초등학교 데이터
        middle = rejs['schoolList03']  # 중학교 데이터
        high = rejs['schoolList04']  # 고등학교 데이터
        special = rejs['schoolList05']  # 특수학교 데이터

        if len(elementary) == 0 and len(middle) == 0 and len(high) == 0 and len(special) == 0:
            return("CAN NOT FIND SCHOOL")

        # 담아서 넘길 리스트 만들기
        self.elementary = []
        self.middle = []
        self.high = []
        self.special = []

        if len(elementary) > 0:  # 초등학교 데이터 분석
            for i in range(0, len(elementary)):
                sinfo = elementary[i]
                sname = sinfo['SCHUL_NM']
                saddress = sinfo['SCHUL_RDNMA']
                sncode = sinfo['SCHUL_CODE']

                sinfo = {
                    'NAME': sname,
                    'ADDRESS': saddress,
                    'CODE': sncode
                }

                self.elementary.append(sinfo)

        if len(middle) > 0:  # 중학교 데이터 분석
            for i in range(0, len(middle)):
                sinfo = middle[i]
                sname = sinfo['SCHUL_NM']
                saddress = sinfo['SCHUL_RDNMA']
                sncode = sinfo['SCHUL_CODE']

                sinfo = {
                    'NAME': sname,
                    'ADDRESS': saddress,
                    'CODE': sncode
                }

                self.middle.append(sinfo)

        if len(high) > 0:  # 고등학교 데이터 분석
            for i in range(0, len(high)):
                sinfo = high[i]
                sname = sinfo['SCHUL_NM']
                saddress = sinfo['SCHUL_RDNMA']
                sncode = sinfo['SCHUL_CODE']

                sinfo = {
                    'NAME': sname,
                    'ADDRESS': saddress,
                    'CODE': sncode
                }

                self.high.append(sinfo)

        if len(special) > 0:  # 특수학교 데이터 분석
            for i in range(0, len(special)):
                sinfo = special[i]
                sname = sinfo['SCHUL_NM']
                saddress = sinfo['SCHUL_RDNMA']
                sncode = sinfo['SCHUL_CODE']

                sinfo = {
                    'NAME': sname,
                    'ADDRESS': saddress,
                    'CODE': sncode
                }

                self.special.append(sinfo)

        if kind == "1":
            slist = self.elementary
        elif kind == "2":
            slist = self.middle
        elif kind == "3":
            slist = self.high
        elif kind == "4":
            slist = self.special
        elif kind == "0":
            slist = [self.elementary, self.middle, self.high, self.special]
        else:
            slist = [self.elementary, self.middle, self.high, self.special]

        return(slist)

 # ~~ 2. 학교 급식을 불러오는 api ~~


class Smeal:  # 클래스
    def __init__(self, ooe, code, sclass):  # 초기화자 메서드 선언 (기본학교정보설정)
        city_dict = {"서울": "sen.go.kr", "부산": "pen.go.kr", "대구": "dge.go.kr",
                     "인천": "ice.go.kr", "광주": "gen.go.kr", "대전": "dje.go.kr",
                     "울산": "use.go.kr", "세종": "sje.go.kr", "경기": "goe.go.kr",
                     "강원": "kwe.go.kr", "충북": "cbe.go.kr", "충남": "cne.go.kr",
                     "전북": "jbe.go.kr", "전남": "jne.go.kr", "경북": "gbe.kr",
                     "경남": "gne.go.kr", "제주": "jje.go.kr"}  # 학교 목록

        self.ooe = city_dict.get(ooe, "nocity")
        self.sclass = sclass  # 교급
        self.code = code  # 학교 고유 코드

    def day(self, yeon, dal, il, kind):  # 하루치 급식을 조회
        if type(self.ooe) != str or type(yeon) != str or type(self.sclass) != str\
                or type(kind) != str or type(self.code) != str or type(dal) != str\
                or type(il) != str:  # 문자열 형식으로 올바르게 받았는지 확인하는 코드
            return("TYPE ERROR")

        if len(yeon) != 4 or len(dal) != 2 or len(il) != 2\
                or len(self.sclass) != 1 or len(kind) != 1:  # 길이가 알맞는지 확인하는 코드
            return("SIZE ERROR")

        if self.ooe == "nocity":
            return("OFFICE ERROR")

        ymd = yeon + "." + dal + "." + il  # 년도 조합

        url = "http://stu." + self.ooe + "/sts_sci_md01_001.do?" +\
            "schulCode=" + self.code +\
            "&schulCrseScCode=" + self.sclass +\
            "&schulKnaScCode=0" + self.sclass +\
            "&schMmealScCode=" + kind +\
            "&schYmd=" + ymd  # 나이스 학교 급식 조회 주소

        response = requests.get(url)  # 급식 정보 조회
        if response.status_code != 200:  # 응답이 200 (정상응답)이 아닐경우
            return('SERVER ERROR')  # 에러 반환

        foodhtml = BeautifulSoup(response.text, 'html.parser')  # 급식정보 파싱 준비
        foodhtml_data_tr = foodhtml.find_all('tr')  # 모든 tr태그 불러오기

        # 몇번째 행에 급식 정보가 존재하는지 구분하는 로직

        foodhtml_data = foodhtml_data_tr[0].find_all('th')  # 날짜 정보가 있는 열을 불러옴

        try:  # 예외 처리를 위한 try
            for i in range(1, 7):  # 월요일부터 일요일까지 하나하나 대입 준비
                date = str(foodhtml_data[i])  # i요일째 날짜 정보 확인
                date_filter = ['<th class="point2" scope="col">', '<th class="last point1" scope="col">',
                               '<th scope="col">', '</th>', '(', ')', '일', '월', '화', '수', '목', '금', '토']  # 제거해야 하는 목록

                for sakje in date_filter:
                    date = date.replace(sakje, '')  # 찌끄레기를 삭제

                if date != ymd:  # 날짜와 입력날짜 동일 여부 확인
                    continue

                hang = i - 1  # 급식정보가 존재하는 행 선언
                break  # 존재 확인 로직 정지

        except:  # 에러 발생시 데이터베이스 에러 반환
            return("NO DATABASE")

        # 급식 정보 조회 시작

        try:
            food = foodhtml_data_tr[2].find_all(
                'td')  # 급식정보가 있는 행의 모든 td 태그 불러오기
            food = str(food[hang])  # hang 번째에 있는 급식 정보 불러옴

            food_filter = ['<td class="textC">',
                           '<td class="textC last">', '</td>']  # 제거해야 하는 목록

            for sakje in food_filter:
                food = food.replace(sakje, '')  # 찌끄레기를 삭제

            if food == ' ':
                food = '급식이 예정되지 않았거나 정보가 존재하지 않습니다.'  # 만약 조회시 급식정보가 없다면 미존재 안내

        except:
            food = 'NO DATABASE'  # 급식 조회 실패시 안내

        return(food)  # 정보 반환

    def month(self, yeon, dal, kind, output):  # 한달치 급식을 조회
        # 조회하는 월의 마지막 날 구하는 로직
        if dal == '02':  # 2월 조회시
            if yeon % 4 == 0 and yeon % 100 != 0:  # 윤년일 경우 29일이 마지막
                last_day = 29
            elif yeon % 400 == 0:
                last_day = 29
            else:  # 아니면 28이 마지막
                last_day = 28
        elif dal == '01' or dal == '03' or dal == '05' or dal == '07' or dal == '08' or dal == '10' or dal == '12':  # 끝날이 31일 목록
            last_day = 31
        else:  # 이외는 모두 30일이 마지막
            last_day = 30

        if output == "e":  # 엑셀 저장 기능
            op = openpyxl.Workbook()
            ex = op.active

            try:  # 예외 처리를 위한 try문
                for i in range(1, last_day + 1):  # 한달치 모두 대입하는 반복문
                    if i < 10:  # 일 정보가 10 미만일때
                        i = str(i)
                        i = "0" + i  # 매개변수 입력 규칙에 의거 두자리로 변환

                    i = str(i)

                    meal = self.day(yeon, dal, i, kind)  # 급식 정보 불러옴
                    ex.cell(row=int(i), column=1).value = yeon + \
                        "년" + dal + "월" + i + "일"  # 날짜 정보 엑셀 삽입
                    ex.cell(row=int(i), column=2).value = meal  # 급식 정보 삽입

                op.save(self.code + "_" + yeon + "년 " + dal + "월" + ".xlsx")
                op.close
                return("SUCCEED")

            except:  # 실패시 실패 에러 안내
                op.close()  # 엑셀 닫기
                return("EXCEL ERROR")

        if output == "c":  # csv 저장 기능
            f = open(self.code + "_" + yeon + "년 " + dal + "월" +
                     ".csv", 'w', encoding='utf-8', newline='')
            cs = csv.writer(f)

            try:  # 예외 처리를 위한 try문
                for i in range(1, last_day + 1):  # 한달치 모두 대입하는 반복문
                    if i < 10:  # 일 정보가 10 미만일때
                        i = str(i)
                        i = "0" + i  # 매개변수 입력 규칙에 의거 두자리로 변환

                    i = str(i)

                    meal = self.day(yeon, dal, i, kind)  # 급식 정보 불러옴
                    cs.writerow([yeon +
                                 "년" + dal + "월" + i + "일", meal])  # csv 저장

                f.close()  # csv 닫기
                return("SUCCEED")

            except:  # 실패시 실패 에러 안내
                f.close()  # csv 닫기
                return("CSV ERROR")
